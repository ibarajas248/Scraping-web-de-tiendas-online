#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


INPUT_XLSX = "formato_1.xlsx"                 # <- tu archivo
OUTPUT_XLSX = "pivot_precios_heatmap1.xlsx"  # <- salida


def to_float(x):
    """Convierte precios a float (soporta números y strings con símbolos/separadores)."""
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x)
    s = re.sub(r"[^\d,.\-]", "", s)

    # Manejo básico de coma/punto
    if s.count(",") > 0 and s.count(".") == 0:
        parts = s.split(",")
        if len(parts[-1]) == 2:
            s = "".join(parts[:-1]) + "." + parts[-1]
        else:
            s = "".join(parts)
    else:
        s = s.replace(",", "")

    try:
        return float(s)
    except:
        return np.nan


def main():
    # 1) Leer archivo
    df = pd.read_excel(INPUT_XLSX)

    # 2) Normalizar columnas clave
    df["EAN"] = pd.to_numeric(df["EAN"], errors="coerce").astype("Int64")
    df = df[df["EAN"].notna()].copy()
    df["EAN"] = df["EAN"].astype(int).astype(str)

    df["PRODUCTO"] = df["PRODUCTO"].astype(str).str.strip()
    df["BANDERA"] = df["BANDERA"].astype(str).str.strip()

    # FECHA para “último registro”
    df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")
    df["_row"] = np.arange(len(df))  # desempate: orden real del archivo

    # 3) Quedarse con el ÚLTIMO por (EAN, tienda)
    df = df.sort_values(["EAN", "BANDERA", "FECHA", "_row"])
    df_last = df.groupby(["EAN", "BANDERA"], as_index=False).tail(1).copy()

    # 4) Precio final: oferta primero; si no hay, precio base (PRECIO_LISTA)
    df_last["PRECIO_OFERTA_F"] = df_last["PRECIO_OFERTA"].apply(to_float)
    df_last["PRECIO_BASE_F"] = df_last["PRECIO_LISTA"].apply(to_float)

    df_last["PRECIO_FINAL"] = np.where(
        df_last["PRECIO_OFERTA_F"].notna() & (df_last["PRECIO_OFERTA_F"] > 0),
        df_last["PRECIO_OFERTA_F"],
        df_last["PRECIO_BASE_F"],
    )

    # 5) Unificar nombre de producto por EAN (toma el más frecuente)
    prod_map = (
        df_last.groupby("EAN")["PRODUCTO"]
        .agg(lambda s: s.value_counts().index[0] if len(s.value_counts()) else s.iloc[0])
    )
    df_last["PRODUCTO_STD"] = df_last["EAN"].map(prod_map)

    # 6) Pivot: filas = (EAN, PRODUCTO), columnas = tiendas, valores = PRECIO_FINAL
    pivot = (
        df_last.pivot_table(
            index=["EAN", "PRODUCTO_STD"],
            columns="BANDERA",
            values="PRECIO_FINAL",
            aggfunc="first",
        )
        .reset_index()
        .rename(columns={"PRODUCTO_STD": "PRODUCTO"})
    )
    # --- NUEVO: ordenar primero los "coincidentes" (EAN presentes en varias tiendas) ---
    tienda_cols = [c for c in pivot.columns if c not in ("EAN", "PRODUCTO")]
    pivot["COINCIDENCIAS"] = pivot[tienda_cols].notna().sum(axis=1)  # cuántas tiendas tienen precio

    # primero los de más coincidencias, luego por EAN (opcional)
    pivot = pivot.sort_values(["COINCIDENCIAS", "EAN"], ascending=[False, True])

    # si NO quieres mostrar la columna COINCIDENCIAS en el excel, bórrala después de ordenar:
    # pivot = pivot.drop(columns=["COINCIDENCIAS"])
    # 7) Exportar a Excel
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        pivot.to_excel(writer, index=False, sheet_name="Pivot")

    # 8) Aplicar formato + heatmap por fila (min→verde, max→rojo)
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["Pivot"]

    # Congelar encabezado + primeras 2 columnas
    ws.freeze_panes = "C2"

    # Encabezado en negrita
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Anchos
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Formato números (miles) + alineación
    for r in range(2, ws.max_row + 1):
        for c in range(3, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # Heatmap por fila (solo columnas de tiendas)
    start_col = 3
    end_col = ws.max_column
    for r in range(2, ws.max_row + 1):
        rng = f"{get_column_letter(start_col)}{r}:{get_column_letter(end_col)}{r}"
        rule = ColorScaleRule(
            start_type="min",
            start_color="63BE7B",      # verde (barato)
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",        # amarillo (medio)
            end_type="max",
            end_color="F8696B",        # rojo (caro)
        )
        ws.conditional_formatting.add(rng, rule)

    # Filtro
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(OUTPUT_XLSX)
    print(f"OK -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
